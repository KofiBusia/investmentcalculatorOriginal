# Rebuild RESIDUAL INCOME VALUATION MODEL.xlsx with three ROE forecast modes.

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.comments import Comment

OUT = r"C:\Users\kkyei\Desktop\RESIDUAL INCOME VALUATION MODEL.xlsx"

# ── colour palette ────────────────────────────────────────────
NAVY   = "0D2B55"
GOLD   = "C9A02C"
TEAL   = "097A6E"
LBLUE  = "EFF6FF"
LGREEN = "F0FDF4"
LYELLOW= "FFFBEB"
WHITE  = "FFFFFF"
LGRAY  = "F8FAFC"
LGRAY2 = "F1F5F9"
RED    = "FEE2E2"

def fill(hex_):   return PatternFill("solid", fgColor=hex_)
def font(bold=False, size=11, color="374151", italic=False):
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)
def bottom_border():
    s = Side(style="thin", color="CBD5E1")
    return Border(bottom=s)

def style(cell, bold=False, size=11, color="374151", bg=None, h="left",
          v="center", wrap=False, italic=False, border=False):
    cell.font      = font(bold=bold, size=size, color=color, italic=italic)
    cell.alignment = align(h=h, v=v, wrap=wrap)
    if bg:   cell.fill   = fill(bg)
    if border: cell.border = thin_border()

def hdr_row(ws, row, text, bg=NAVY, color=WHITE, size=13, cols=12):
    c = ws.cell(row=row, column=1, value=text)
    style(c, bold=True, size=size, color=color, bg=bg, h="left", v="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 22

def label(ws, row, col, text, bold=False, bg=None, color="374151",
          wrap=False, size=11, italic=False, h="left"):
    c = ws.cell(row=row, column=col, value=text)
    style(c, bold=bold, size=size, color=color, bg=bg, wrap=wrap, italic=italic, h=h)
    return c

def inp(ws, row, col, value=None, fmt=None, bg=LYELLOW, border=True, h="right"):
    c = ws.cell(row=row, column=col, value=value)
    style(c, bold=True, size=11, color=NAVY, bg=bg, h=h, border=border)
    if fmt: c.number_format = fmt
    return c

def formula(ws, row, col, expr, fmt=None, bg=LGRAY, bold=False, color="374151"):
    c = ws.cell(row=row, column=col, value=expr)
    style(c, bold=bold, size=11, color=color, bg=bg, h="right")
    if fmt: c.number_format = fmt
    return c

wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ══════════════════════════════════════════════════════════════
cover = wb.active
cover.title = "Cover"
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 32
cover.column_dimensions["B"].width = 38

hdr_row(cover, 1, "  BANK STOCK INTRINSIC VALUE CALCULATOR", bg=NAVY, size=15, cols=2)
hdr_row(cover, 2, "  Residual Income Model — Dividend-Irregularity Adjusted", bg=TEAL, size=12, cols=2)

data = [
    (4,  "Model",            "Residual Income Model (RIM)"),
    (5,  "Core Formula",     "IV = BV₀ + Σ[(ROE−r)×BVₜ₋₁/(1+r)ᵗ] + TV/(1+r)ⁿ"),
    (6,  "Cost of Equity",   "CAPM: r = Rf + β × (Rm − Rf)"),
    (7,  "ROE Forecast Modes","Manual | Auto from History | Growth Rate"),
    (8,  "Dividend Policy",  "No Dividends / Regular / Irregular (with irregularity factor)"),
    (10, "Intrinsic Value",  "=IF(Valuation!B14=0,\"Awaiting inputs\",Valuation!B14)"),
    (11, "Market Price",     "=IF(Inputs!B28=0,\"Not entered\",Inputs!B28)"),
    (12, "Margin of Safety", "=IFERROR(IF(Valuation!B14=0,\"—\",(Valuation!B14-Inputs!B28)/Valuation!B14),\"—\")"),
    (13, "Signal",           "=IFERROR(IF(Valuation!B14=0,\"—\",IF(Valuation!B14>Inputs!B29,\"UNDERVALUED\",IF(Valuation!B14<Inputs!B29,\"OVERVALUED\",\"FAIRLY VALUED\"))),\"—\")"),
]
for row, lbl, val in data:
    c1 = cover.cell(row=row, column=1, value=lbl)
    style(c1, bold=True, size=11, color=NAVY, bg=LGRAY2)
    c2 = cover.cell(row=row, column=2, value=val)
    style(c2, size=11, color="374151", bg=WHITE)

cover.cell(row=15, column=1, value="Note").font = font(italic=True, color="9CA3AF")
note = cover.cell(row=16, column=1, value="Enter data only on the Inputs sheet. Use the ROE_Engine sheet to set up your ROE forecast method.")
style(note, italic=True, size=10, color="9CA3AF", bg=WHITE, wrap=True)
cover.merge_cells("A16:B16")
cover.row_dimensions[16].height = 32

# ══════════════════════════════════════════════════════════════
# SHEET 2 — ROE_ENGINE  (the new comprehensive sheet)
# ══════════════════════════════════════════════════════════════
roe_eng = wb.create_sheet("ROE_Engine")
roe_eng.sheet_view.showGridLines = False
for col, w in [(1,30),(2,14),(3,14),(4,14),(5,14),(6,14),(7,20)]:
    roe_eng.column_dimensions[get_column_letter(col)].width = w

hdr_row(roe_eng, 1, "  ROE FORECAST ENGINE", bg=NAVY, size=14, cols=7)

# Mode selector
label(roe_eng, 3, 1, "ROE Forecast Mode", bold=True, size=11, color=NAVY)
mode_cell = inp(roe_eng, 3, 2, value="Manual", bg=LYELLOW)
dv = DataValidation(type="list", formula1='"Manual,Auto-Average,Auto-CAGR,Growth Rate,From-EPS"', allow_blank=False)
roe_eng.add_data_validation(dv)
dv.add(mode_cell)
label(roe_eng, 3, 3, "↑ Manual | Auto-Average | Auto-CAGR | Growth Rate | From-EPS", italic=True, size=9, color="9CA3AF")
roe_eng.merge_cells("C3:G3")

# ── Section A: Historical ROE ─────────────────────────────────
hdr_row(roe_eng, 5, "  SECTION A — Historical ROE (needed for Auto-Average and Auto-CAGR modes)", bg=TEAL, size=11, cols=7)
years_hist = ["Y-5 (oldest)","Y-4","Y-3","Y-2","Y-1 (latest)"]
for i, lbl in enumerate(years_hist):
    label(roe_eng, 6, i+2, lbl, bold=True, size=10, color=NAVY, bg=LGRAY2, h="center")
label(roe_eng, 7, 1, "Historical ROE (%)", bold=False, size=11)
for i in range(5):
    inp(roe_eng, 7, i+2, value=0, fmt="0.00%")

# Computed stats
label(roe_eng, 9, 1, "Simple Average of Historical ROEs →", italic=True, size=10, color=TEAL)
formula(roe_eng, 9, 2, "=IFERROR(AVERAGE(B7:F7),0)", fmt="0.00%", bold=True, color=TEAL)
label(roe_eng, 10, 1, "Latest ROE (Y-1) →", italic=True, size=10, color=TEAL)
formula(roe_eng, 10, 2, "=F7", fmt="0.00%", bold=True, color=TEAL)
label(roe_eng, 11, 1, "CAGR of ROE (Y-5 to Y-1) →", italic=True, size=10, color=TEAL)
formula(roe_eng, 11, 2, "=IFERROR(POWER(F7/B7,1/4)-1,0)", fmt="0.00%", bold=True, color=TEAL)
label(roe_eng, 11, 3, "(last/first)^(1/4)−1  — adjust denominator if fewer years used", italic=True, size=9, color="9CA3AF")
roe_eng.merge_cells("C11:G11")

# ── Section B: Growth Rate mode ───────────────────────────────
hdr_row(roe_eng, 13, "  SECTION B — Growth Rate Mode", bg=GOLD, color=NAVY, size=11, cols=7)
label(roe_eng, 14, 1, "Base ROE — latest known (%)", bold=False)
inp(roe_eng, 14, 2, value=0, fmt="0.00%")
label(roe_eng, 15, 1, "Annual ROE Growth Rate (%)")
inp(roe_eng, 15, 2, value=0, fmt="0.00%")

# ── Section C: Manual override ────────────────────────────────
hdr_row(roe_eng, 17, "  SECTION C — Manual ROE Forecast Override", bg=LGRAY2, color=NAVY, size=11, cols=7)
forecast_yrs = ["Year 1","Year 2","Year 3","Year 4","Year 5"]
for i, lbl in enumerate(forecast_yrs):
    label(roe_eng, 18, i+2, lbl, bold=True, size=10, color=NAVY, bg=LGRAY2, h="center")
label(roe_eng, 19, 1, "Manual ROE Forecast (%)", bold=False)
for i in range(5):
    inp(roe_eng, 19, i+2, value=0, fmt="0.00%")

# ── Section D: ACTIVE forecast (the one the model uses) ───────
hdr_row(roe_eng, 21, "  ACTIVE FORECAST — used by the model (auto-selected based on Mode)", bg=NAVY, size=11, cols=7)
label(roe_eng, 22, 1, "Selected ROE Forecast (%)", bold=True, color=NAVY)

# Formula: picks correct ROE based on mode (now includes From-EPS pulling from Section E row 37)
for i in range(5):
    col = i + 2
    col_let = get_column_letter(col)
    yr = i + 1
    f = (f'=IF($B$3="Manual",{col_let}19,'
         f'IF($B$3="Auto-Average",$B$9,'
         f'IF($B$3="Auto-CAGR",IFERROR($F$7*POWER(1+$B$11,{yr}),0),'
         f'IF($B$3="Growth Rate",IFERROR($B$14*POWER(1+$B$15,{yr}),0),'
         f'IF($B$3="From-EPS",IFERROR({col_let}37,0),0)))))')
    c = formula(roe_eng, 22, col, f, fmt="0.00%", bold=True, color=WHITE)
    c.fill = fill(NAVY)
    c.font = Font(bold=True, size=11, color=WHITE, name="Calibri")

label(roe_eng, 23, 1, "↑ Feeds Inputs sheet ROE rows. 'From-EPS' mode uses Section E row 37 (EPS ÷ Rolling BV).", italic=True, size=9, color="9CA3AF")
roe_eng.merge_cells("A23:G23")

# ══════════════════════════════════════════════════════════════
# ROE_Engine — SECTION E: ROE FROM FINANCIAL STATEMENTS
# ══════════════════════════════════════════════════════════════
hdr_row(roe_eng, 25, "  SECTION E — CALCULATE ROE FROM FINANCIAL STATEMENTS", bg="5B21B6", size=12, cols=7)

# Sub-section E1: Historical NI + Equity → historical ROE
label(roe_eng, 26, 1, "E1 — Historical Financial Data (up to 5 years from Annual Reports)", bold=True, size=11, color="5B21B6")
label(roe_eng, 26, 3, "Use CONSISTENT units: all in GHS thousands OR all in GHS millions.", italic=True, size=9, color="9CA3AF")
roe_eng.merge_cells("C26:G26")

years_hist_e = ["Y-5 (oldest)", "Y-4", "Y-3", "Y-2", "Y-1 (latest)"]
for i, lbl in enumerate(years_hist_e):
    label(roe_eng, 27, i+2, lbl, bold=True, size=9, color="5B21B6", bg="F5F3FF", h="center")

# NI row
label(roe_eng, 28, 1, "Net Income  [Income Statement — 'Profit for the Year' / 'Net Profit After Tax']", bold=False, size=10, wrap=True)
for i in range(5):
    inp(roe_eng, 28, i+2, value=None, fmt="#,##0", bg=LYELLOW)
roe_eng.row_dimensions[28].height = 32

# Equity row
label(roe_eng, 29, 1, "Shareholders' Equity  [Balance Sheet — 'Total Equity' — use beginning-of-year balance]", bold=False, size=10, wrap=True)
for i in range(5):
    inp(roe_eng, 29, i+2, value=None, fmt="#,##0", bg=LYELLOW)
roe_eng.row_dimensions[29].height = 32

# Computed historical ROE = NI / Equity
label(roe_eng, 30, 1, "Computed Historical ROE = NI ÷ Equity  (auto)", bold=True, size=10, color="5B21B6")
for i in range(5):
    col_let = get_column_letter(i+2)
    c = roe_eng.cell(row=30, column=i+2, value=f'=IFERROR({col_let}28/{col_let}29,"")')
    style(c, bold=True, size=11, color="5B21B6", bg="EDE9FE", h="right")
    c.number_format = "0.00%"

label(roe_eng, 31, 1, "NEXT STEP: Change Mode (B3) to 'Auto-Average' or 'Auto-CAGR', then copy row 30 values into Section A row 7 to use these historical ROEs.", italic=True, size=9, color="9CA3AF", wrap=True)
roe_eng.merge_cells("A31:G31")
roe_eng.row_dimensions[31].height = 30

# Sub-section E2: EPS-based Forecast ROE (for "From-EPS" mode)
label(roe_eng, 33, 1, "E2 — Forecast ROE from Projected EPS  (set Mode = 'From-EPS' to activate)", bold=True, size=11, color="92400E")
label(roe_eng, 33, 3, "ROE = EPS ÷ Opening BV per Share. BV rolls forward each year via retained earnings.", italic=True, size=9, color="9CA3AF")
roe_eng.merge_cells("C33:G33")

for i, lbl in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    label(roe_eng, 34, i+2, lbl, bold=True, size=9, color="92400E", bg="FFFBEB", h="center")

# Projected EPS row
label(roe_eng, 35, 1, "Projected EPS  [Analyst forecast OR: Net Income ÷ Shares Outstanding]", bold=False, size=10)
for i in range(5):
    inp(roe_eng, 35, i+2, value=None, fmt="0.0000", bg=LYELLOW)

# Opening BV per share (rolling forward)
label(roe_eng, 36, 1, "Opening BV per share (rolling — auto-calculated from BV0 + retained earnings)", bold=True, size=10, color="92400E")
# Year 1 opening BV from Inputs sheet
c_bv1 = roe_eng.cell(row=36, column=2, value="=IFERROR(Inputs!B15,0)")
style(c_bv1, bold=True, size=10, color="92400E", bg="FEF9C3", h="right")
c_bv1.number_format = "0.0000"
for i in range(1, 5):  # Year 2..5
    prev_col = get_column_letter(i+1)
    roll_f = f"=IFERROR({prev_col}36+{prev_col}35*(1-Inputs!$B$19),0)"
    c = roe_eng.cell(row=36, column=i+2, value=roll_f)
    style(c, bold=True, size=10, color="92400E", bg="FEF9C3", h="right")
    c.number_format = "0.0000"

# Computed Forecast ROE = EPS / Opening BV → referenced by Active Forecast "From-EPS"
label(roe_eng, 37, 1, "Computed Forecast ROE = EPS ÷ Opening BV  [Active when Mode = 'From-EPS']", bold=True, size=10, color="5B21B6")
for i in range(5):
    col_let = get_column_letter(i+2)
    c = roe_eng.cell(row=37, column=i+2, value=f'=IFERROR({col_let}35/{col_let}36,"")')
    style(c, bold=True, size=11, color="5B21B6", bg="EDE9FE", h="right")
    c.number_format = "0.00%"
    c.font = Font(bold=True, size=11, color="5B21B6", name="Calibri")

label(roe_eng, 38, 1, "Set Mode dropdown (B3) to 'From-EPS' → row 37 values auto-flow into the Active Forecast row 22 above.", italic=True, size=9, color="9CA3AF", wrap=True)
roe_eng.merge_cells("A38:G38")
roe_eng.row_dimensions[38].height = 28

# ══════════════════════════════════════════════════════════════
# ROE_Engine — SECTION F: WHERE TO FIND EACH INPUT
# ══════════════════════════════════════════════════════════════
hdr_row(roe_eng, 40, "  SECTION F — FIELD GUIDE: WHERE TO FIND EACH INPUT IN FINANCIAL STATEMENTS", bg=TEAL, size=11, cols=7)

guidance = [
    ("INPUT",                   "WHERE TO FIND IT",                                                              "TYPICAL VALUE (Ghana)",    TEAL,    WHITE,   WHITE),
    ("Net Income",              "Income Statement — last line: 'Profit for the Year' or 'Net Profit After Tax'", "GHS millions",             "F0FDF4", NAVY,   "16A34A"),
    ("Shareholders' Equity",    "Balance Sheet — 'Total Equity' (use prior year closing balance as opening BV)",  "GHS millions",             "F0FDF4", NAVY,   "16A34A"),
    ("BV per Share",            "Balance Sheet Total Equity ÷ Shares Outstanding   OR   GSE Fact Sheet",          "e.g. GHS 4.20/share",      "EFF6FF", NAVY,   "4F46E5"),
    ("EPS",                     "Income Statement (per-share note)   OR   Net Income ÷ Shares in Issue",          "e.g. GHS 0.75/share",      "EFF6FF", NAVY,   "4F46E5"),
    ("Payout Ratio",            "Annual Report: Directors' Report / Dividend Announcement.  DPS ÷ EPS × 100",     "30–60% typical",           "FFFBEB", NAVY,   GOLD),
    ("Beta (β)",                "Bloomberg / Reuters / Calculated: Cov(stock,index) ÷ Var(index) over 3–5 yrs",  "0.5–1.2 for GH banks",    "FFFBEB", NAVY,   GOLD),
    ("Risk-Free Rate",          "Bank of Ghana (bog.gov.gh) → T-bill rates / Bond yields page",                   "~25–30% (2024–25)",        "FEF9C3", NAVY,   "92400E"),
    ("Market Return (Rm)",      "GSE website (gse.com.gh): GSE All-Share Index annualised 5–10 yr return",        "~20–28% (GSE historical)", "FEF9C3", NAVY,   "92400E"),
]
for i, (inp_name, where, example, bg_, text_, hint_) in enumerate(guidance):
    row = 41 + i
    c1 = roe_eng.cell(row=row, column=1, value=inp_name)
    style(c1, bold=(i==0), size=10 if i>0 else 11, color=text_ if i==0 else hint_, bg=bg_, wrap=(i>0))
    c2 = roe_eng.cell(row=row, column=2, value=where)
    style(c2, bold=False, size=9 if i>0 else 11, color=text_ if i==0 else "374151", bg=bg_, wrap=True)
    roe_eng.merge_cells(f"B{row}:F{row}")
    c3 = roe_eng.cell(row=row, column=7, value=example)
    style(c3, bold=False, size=9 if i>0 else 11, color=text_ if i==0 else "6B7280", bg=bg_, italic=(i>0), h="center")
    roe_eng.row_dimensions[row].height = 22 if i==0 else 20

# ══════════════════════════════════════════════════════════════
# SHEET 3 — INPUTS
# ══════════════════════════════════════════════════════════════
inp_ws = wb.create_sheet("Inputs")
inp_ws.sheet_view.showGridLines = False
for col, w in [(1,38),(2,18),(3,18),(4,18),(5,18),(6,18)]:
    inp_ws.column_dimensions[get_column_letter(col)].width = w

hdr_row(inp_ws, 1, "  INPUTS & ASSUMPTIONS", bg=NAVY, size=14, cols=6)

# Setup
hdr_row(inp_ws, 3, "  1. SETUP", bg=TEAL, size=11, cols=6)
rows_setup = [
    (4,  "Currency",                         "GHS",    None,   False),
    (5,  "Number of forecast years (3–5)",   5,        "0",    True),
]
for row, lbl, val, fmt_, editable in rows_setup:
    label(inp_ws, row, 1, lbl)
    c = inp_ws.cell(row=row, column=2, value=val)
    style(c, bold=True, size=11, color=NAVY, bg=LYELLOW if editable else LGRAY, h="right")
    if fmt_: c.number_format = fmt_

dv_years = DataValidation(type="whole", operator="between", formula1="3", formula2="5")
inp_ws.add_data_validation(dv_years)
dv_years.add(inp_ws["B5"])

# CAPM
hdr_row(inp_ws, 7, "  2. COST OF EQUITY — CAPM  [r = Rf + β × (Rm − Rf)]", bg=TEAL, size=11, cols=6)
capm_rows = [
    (8,  "Risk-free rate — long-term govt bond yield (%)", 0,   "0.00%", True),
    (9,  "Expected market return (%)",                     0,   "0.00%", True),
    (10, "Beta (stock volatility vs market)",              0,   "0.00",  True),
    (11, "Market risk premium (MRP = Rm − Rf)",           "=B9-B8", "0.00%", False),
    (12, "Cost of Equity, r = Rf + β × MRP",              "=B8+B10*(B9-B8)", "0.00%", False),
]
for row, lbl, val, fmt_, editable in capm_rows:
    label(inp_ws, row, 1, lbl, bold=not editable)
    c = inp_ws.cell(row=row, column=2, value=val)
    style(c, bold=True, size=11, color=NAVY if editable else TEAL,
          bg=LYELLOW if editable else LGRAY, h="right")
    c.number_format = fmt_

# RIM Inputs
hdr_row(inp_ws, 14, "  3. RESIDUAL INCOME INPUTS", bg=TEAL, size=11, cols=6)
rim_rows = [
    (15, "Current Book Value per Share (BV₀)",    0,   "0.00",  True),
    (16, "Dividend Policy",                       "Regular Dividends", None, True),
    (17, "Dividend Payout Ratio — stated (%)",    0,   "0.00%", True),
    (18, "Irregularity Factor (0–1; for Irregular policy)", 1, "0.00", True),
    (19, "Effective Payout used in model",        "=IF(B16=\"No Dividends\",0,IF(B16=\"Irregular Dividends\",B17*B18,B17))", "0.00%", False),
]
for row, lbl, val, fmt_, editable in rim_rows:
    label(inp_ws, row, 1, lbl, bold=not editable)
    c = inp_ws.cell(row=row, column=2, value=val)
    style(c, bold=True, size=11, color=NAVY if editable else TEAL,
          bg=LYELLOW if editable else LGRAY, h="right")
    if fmt_: c.number_format = fmt_

dv_pol = DataValidation(type="list", formula1='"No Dividends,Regular Dividends,Irregular Dividends"')
inp_ws.add_data_validation(dv_pol)
dv_pol.add(inp_ws["B16"])
label(inp_ws, 18, 3, "1.0 = dividends paid every year; 0.75 = 3 out of 4 years", italic=True, size=9, color="9CA3AF")
inp_ws.merge_cells("C18:F18")

# Terminal Value
hdr_row(inp_ws, 21, "  4. TERMINAL VALUE ASSUMPTIONS", bg=TEAL, size=11, cols=6)
tv_rows = [
    (22, "Terminal growth rate, g (%)",          0,   "0.00%", True),
    (23, "Long-term sustainable ROE (%)",         0,   "0.00%", True),
    (24, "CHECK: LT ROE > Cost of Equity?",      '=IF(B23>B12,"✅ YES — value-creating","❌ NO — destroys value")', None, False),
    (25, "CHECK: g < Cost of Equity?",           '=IF(B22<B12,"✅ YES — model valid","❌ NO — fix g or CAPM inputs")', None, False),
]
for row, lbl, val, fmt_, editable in tv_rows:
    label(inp_ws, row, 1, lbl, bold=not editable)
    c = inp_ws.cell(row=row, column=2, value=val)
    style(c, bold=True, size=11, color=NAVY if editable else TEAL,
          bg=LYELLOW if editable else LGRAY, h="right" if editable else "left")
    if fmt_: c.number_format = fmt_

# Market Price
hdr_row(inp_ws, 27, "  5. MARKET COMPARISON (optional)", bg=TEAL, size=11, cols=6)
label(inp_ws, 28, 1, "Current market price per share")
inp(inp_ws, 28, 2, value=0, fmt="0.00")

# ROE Forecast — links to ROE_Engine
hdr_row(inp_ws, 30, "  6. ROE FORECAST BY YEAR — linked to ROE_Engine (change mode there)", bg=GOLD, color=NAVY, size=11, cols=6)
yr_labels = ["Year 1","Year 2","Year 3","Year 4","Year 5"]
for i, lbl in enumerate(yr_labels):
    label(inp_ws, 31, i+2, lbl, bold=True, size=10, color=NAVY, bg=LGRAY2, h="center")
label(inp_ws, 32, 1, "Active Forecasted ROE", bold=True, color=NAVY)
for i in range(5):
    col_let = get_column_letter(i+2)
    eng_col = get_column_letter(i+2)
    c = formula(inp_ws, 32, i+2, f"=ROE_Engine!{eng_col}22", fmt="0.00%", bold=True, color=NAVY)
    c.fill = fill(LBLUE)

label(inp_ws, 33, 1, "Override: Manual ROE entry here (type directly to override ROE_Engine link)", italic=True, size=9, color="9CA3AF")
inp_ws.merge_cells("A33:F33")

# ══════════════════════════════════════════════════════════════
# SHEET 4 — FORECAST
# ══════════════════════════════════════════════════════════════
fc = wb.create_sheet("Forecast")
fc.sheet_view.showGridLines = False
for col, w in [(1,38),(2,16),(3,16),(4,16),(5,16),(6,16)]:
    fc.column_dimensions[get_column_letter(col)].width = w

hdr_row(fc, 1, "  FORECAST ENGINE — Book Value Roll-Forward & Residual Income", bg=NAVY, size=13, cols=6)
label(fc, 2, 1, "All inputs drawn from Inputs sheet. Years beyond forecast horizon are hidden.", italic=True, size=9, color="9CA3AF")

for i in range(5):
    col = i+2
    c = fc.cell(row=3, column=col, value=f"Year {i+1}")
    style(c, bold=True, size=10, color=WHITE, bg=NAVY, h="center")

fc_rows = [
    (4,  "Active in forecast?",             [f'=IF({i+1}<=Inputs!$B$5,"Yes","—")' for i in range(5)],     None,    False),
    (5,  "Forecasted ROE (%)",              [f"=IF({i+1}<=Inputs!$B$5,Inputs!{get_column_letter(i+2)}32,\"\")" for i in range(5)], "0.00%", False),
    (6,  "Opening Book Value per share",    None, "0.00", False),
    (7,  "Net Income per share (ROE×BV)",   None, "0.00", False),
    (8,  "Dividend per share (eff.payout)", None, "0.00", False),
    (9,  "Retained Earnings per share",     None, "0.00", False),
    (10, "Closing Book Value per share",    None, "0.00", False),
    (11, "Equity Charge (r × opening BV)", None, "0.00", False),
    (12, "Residual Income per share",       None, "0.00", False),
    (13, "Discount factor  1/(1+r)ᵗ",      None, "0.0000", False),
    (14, "PV of Residual Income",           None, "0.00", False),
]

label_bgs = {4:LGRAY2, 5:LBLUE, 6:LGRAY, 7:LGRAY, 8:LGRAY, 9:LGRAY, 10:LGRAY2, 11:LGRAY, 12:LGREEN, 13:LGRAY, 14:LGREEN}
for row, lbl, formulas_, fmt_, _ in fc_rows:
    label(fc, row, 1, lbl, bold=(row in [6,10,12,14]), size=10)
    for i in range(5):
        col = i+2
        if formulas_:
            f = formulas_[i]
        else:
            col_let = get_column_letter(col)
            prev_col = get_column_letter(col-1)
            r = row
            if r == 6:   # Opening BV
                f = (f'=IF({i+1}>Inputs!$B$5,"",IF({i+1}=1,Inputs!$B$15,{prev_col}10))' if i > 0
                     else '=IF(1<=Inputs!$B$5,Inputs!$B$15,"")')
            elif r == 7: f = f'=IF({col_let}4="—","",{col_let}5*{col_let}6)'
            elif r == 8: f = f'=IF({col_let}4="—","",{col_let}7*Inputs!$B$19)'
            elif r == 9: f = f'=IF({col_let}4="—","",{col_let}7-{col_let}8)'
            elif r == 10:f = f'=IF({col_let}4="—","",{col_let}6+{col_let}9)'
            elif r == 11:f = f'=IF({col_let}4="—","",Inputs!$B$12*{col_let}6)'
            elif r == 12:f = f'=IF({col_let}4="—","",{col_let}7-{col_let}11)'
            elif r == 13:f = f'=IF({col_let}4="—","",1/POWER(1+Inputs!$B$12,{i+1}))'
            elif r == 14:f = f'=IF({col_let}4="—","",{col_let}12*{col_let}13)'
            else: f = ""

        c = fc.cell(row=row, column=col, value=f)
        bg = label_bgs.get(row, LGRAY)
        style(c, size=10, color="374151", bg=bg, h="right")
        if fmt_: c.number_format = fmt_

# ══════════════════════════════════════════════════════════════
# SHEET 5 — VALUATION
# ══════════════════════════════════════════════════════════════
val = wb.create_sheet("Valuation")
val.sheet_view.showGridLines = False
for col, w in [(1,40),(2,20),(3,20)]:
    val.column_dimensions[get_column_letter(col)].width = w

hdr_row(val, 1, "  VALUATION SUMMARY", bg=NAVY, size=14, cols=3)
label(val, 2, 1, "IV = BV₀ + Σ PV(Residual Income) + PV(Terminal Value)", bold=False, italic=True, size=10, color="9CA3AF")

label(val, 4, 1, "Cost of Equity, r",              bold=True, color=NAVY)
formula(val, 4, 2, "=Inputs!B12", fmt="0.00%", bold=True, color=TEAL)
label(val, 5, 1, "Forecast horizon (years)",        bold=True, color=NAVY)
formula(val, 5, 2, "=Inputs!B5", fmt="0", bold=True, color=TEAL)

hdr_row(val, 7, "  Build-up of Intrinsic Value", bg=TEAL, size=11, cols=3)
bup = [
    (8,  "Current Book Value per share (BV₀)",        "=Inputs!B15",  "0.00"),
    (9,  "PV of Residual Income — forecast period",   "=IFERROR(SUM(Forecast!B14:F14),0)", "0.00"),
    (10, "Terminal BV (end of final forecast year)",  "=IFERROR(CHOOSE(Inputs!B5,Forecast!B10,Forecast!C10,Forecast!D10,Forecast!E10,Forecast!F10),0)", "0.00"),
    (11, "Terminal Residual Income (first perp. year)","=IFERROR((Inputs!B23-Inputs!B12)*B10,0)", "0.00"),
    (12, "Terminal Value = RI/(r−g)",                 "=IFERROR(B11*(1+Inputs!B22)/(Inputs!B12-Inputs!B22),0)", "0.00"),
    (13, "PV of Terminal Value",                      "=IFERROR(B12/POWER(1+Inputs!B12,Inputs!B5),0)", "0.00"),
    (14, "INTRINSIC VALUE PER SHARE",                 "=IFERROR(B8+B9+B13,0)", "0.00"),
]
for row, lbl, fml, fmt_ in bup:
    is_iv = (row == 14)
    label(val, row, 1, lbl, bold=is_iv, size=11 if is_iv else 10,
          color=NAVY if is_iv else "374151", bg=LBLUE if is_iv else None)
    c = val.cell(row=row, column=2, value=fml)
    style(c, bold=is_iv, size=13 if is_iv else 11, color=TEAL,
          bg=LBLUE if is_iv else LGRAY, h="right")
    c.number_format = fmt_
    if is_iv:
        val.row_dimensions[row].height = 24

hdr_row(val, 16, "  Market Comparison", bg=TEAL, size=11, cols=3)
comp = [
    (17, "Current market price per share",          "=Inputs!B28",                                          "0.00"),
    (18, "Upside / (downside) per share",           "=IFERROR(B14-B17,\"—\")",                              "0.00"),
    (19, "Margin of Safety (% of intrinsic value)", "=IFERROR((B14-B17)/B14,\"—\")",                        "0.00%"),
    (20, "Price / Book Value",                      "=IFERROR(B17/Inputs!B15,\"—\")",                       "0.00x"),
    (21, "Intrinsic / Book Value",                  "=IFERROR(B14/Inputs!B15,\"—\")",                       "0.00x"),
    (22, "Signal",                                  '=IFERROR(IF(B14=0,"—",IF(B14>B17,"UNDERVALUED","OVERVALUED")),"—")', None),
]
for row, lbl, fml, fmt_ in comp:
    label(val, row, 1, lbl, size=10)
    c = val.cell(row=row, column=2, value=fml)
    style(c, bold=(row==22), size=11, color=TEAL, bg=LGRAY, h="right")
    if fmt_: c.number_format = fmt_

# ══════════════════════════════════════════════════════════════
# SHEET 6 — SENSITIVITY
# ══════════════════════════════════════════════════════════════
sens = wb.create_sheet("Sensitivity")
sens.sheet_view.showGridLines = False
for col, w in [(1,4),(2,22),(3,14),(4,14),(5,14),(6,14),(7,14),(8,14)]:
    sens.column_dimensions[get_column_letter(col)].width = w

hdr_row(sens, 1, "  SENSITIVITY ANALYSIS — Intrinsic Value per Share", bg=NAVY, size=13, cols=8)

# Table A: Terminal Growth (rows) × LT ROE (cols)
label(sens, 3, 2, "TABLE A  —  Terminal Growth Rate (rows)  ×  Long-Term ROE (cols)", bold=True, color=NAVY, size=11)
sens.merge_cells("B3:H3")

lt_roe_offsets = [-0.04, -0.02, 0, 0.02, 0.04]
g_offsets      = [-0.02, -0.01, 0, 0.01, 0.02]

label(sens, 4, 2, "g \\ LT ROE →", bold=True, bg=LGRAY2, size=10, h="center")
for j, d in enumerate(lt_roe_offsets):
    c = sens.cell(row=4, column=3+j, value=f"=TEXT(Inputs!B23+{d},\"0.0%\")")
    style(c, bold=True, size=10, color=WHITE, bg=NAVY, h="center")

for i, gd in enumerate(g_offsets):
    row = 5+i
    gc = sens.cell(row=row, column=2, value=f"=TEXT(Inputs!B22+{gd},\"0.0%\")")
    style(gc, bold=True, size=10, color=NAVY, bg=LGRAY2, h="center")
    for j, rd in enumerate(lt_roe_offsets):
        g_expr   = f"(Inputs!B22+{gd})"
        roe_expr = f"(Inputs!B23+{rd})"
        r_expr   = "Inputs!B12"
        # IV = BV0 + PV_RI + PV_TV; approximate: use Valuation sheet RI and recalc TV only
        # Full formula: BV0 + SUM(PV_RI_forecast) + ((LT_ROE-r)*terminal_BV*(1+g))/(r-g)/(1+r)^n
        f = (f"=IFERROR(IF({g_expr}>={r_expr},\"N/A\","
             f"IF({roe_expr}<={r_expr},\"N/A\","
             f"Inputs!B15+SUM(Forecast!B14:F14)+"
             f"(({roe_expr}-{r_expr})*Valuation!B10*(1+{g_expr}))/({r_expr}-{g_expr})"
             f"/POWER(1+{r_expr},Inputs!B5))),\"N/A\")")
        c = sens.cell(row=row, column=3+j, value=f)
        style(c, size=10, color="374151", bg=WHITE, h="right", border=True)
        c.number_format = "0.00"

# Table B: Beta (rows) × Terminal Growth (cols)
label(sens, 12, 2, "TABLE B  —  Beta (rows)  ×  Terminal Growth Rate (cols)", bold=True, color=NAVY, size=11)
sens.merge_cells("B12:H12")

beta_offsets = [-0.4,-0.2,0,0.2,0.4]
label(sens, 13, 2, "β \\ g →", bold=True, bg=LGRAY2, size=10, h="center")
for j, d in enumerate(g_offsets):
    c = sens.cell(row=13, column=3+j, value=f"=TEXT(Inputs!B22+{d},\"0.0%\")")
    style(c, bold=True, size=10, color=WHITE, bg=GOLD, h="center")
    c.font = Font(bold=True, size=10, color=NAVY, name="Calibri")

for i, bd in enumerate(beta_offsets):
    row = 14+i
    beta_val = f"MAX(0.01,Inputs!B10+{bd})"
    r_new    = f"(Inputs!B8+({beta_val})*(Inputs!B9-Inputs!B8))"
    brow = sens.cell(row=row, column=2, value=f"=TEXT(MAX(0.01,Inputs!B10+{bd}),\"0.00\")")
    style(brow, bold=True, size=10, color=NAVY, bg=LGRAY2, h="center")
    for j, gd in enumerate(g_offsets):
        g_expr = f"(Inputs!B22+{gd})"
        f = (f"=IFERROR(IF({g_expr}>={r_new},\"N/A\","
             f"IF(Inputs!B23<={r_new},\"N/A\","
             f"Inputs!B15+SUM(Forecast!B14:F14)+"
             f"((Inputs!B23-{r_new})*Valuation!B10*(1+{g_expr}))/({r_new}-{g_expr})"
             f"/POWER(1+{r_new},Inputs!B5))),\"N/A\")")
        c = sens.cell(row=row, column=3+j, value=f)
        style(c, size=10, color="374151", bg=WHITE, h="right", border=True)
        c.number_format = "0.00"

# ══════════════════════════════════════════════════════════════
# SHEET 7 — INSTRUCTIONS
# ══════════════════════════════════════════════════════════════
instr = wb.create_sheet("Instructions")
instr.sheet_view.showGridLines = False
instr.column_dimensions["A"].width = 10
instr.column_dimensions["B"].width = 70

hdr_row(instr, 1, "  HOW TO USE THIS MODEL", bg=NAVY, size=14, cols=2)

steps = [
    ("Step 1", "Select ROE Forecast Mode (ROE_Engine sheet)",
     "Open the ROE_Engine sheet and select your preferred mode from the dropdown in cell B3:\n"
     "• Manual — type each year's ROE directly into the Inputs sheet (row 32)\n"
     "• Auto-Average — enter 3–5 historical ROEs; model uses their simple average\n"
     "• Auto-CAGR — enter historical ROEs; model projects using the CAGR trend\n"
     "• Growth Rate — enter a base ROE and annual growth rate; model compounds forward"),
    ("Step 2", "Enter CAPM Inputs (Inputs sheet)",
     "Risk-free rate: yield on 5–10 year government bonds (e.g. Ghana 10yr bond)\n"
     "Market return: expected annual return of the stock market index\n"
     "Beta: stock's volatility relative to the market (from Bloomberg, Reuters, or Yahoo Finance)\n"
     "Cost of Equity is calculated automatically."),
    ("Step 3", "Enter Residual Income Inputs (Inputs sheet)",
     "Book Value per Share: Total Equity ÷ Shares Outstanding (from the latest Balance Sheet)\n"
     "Dividend Policy: No Dividends / Regular / Irregular\n"
     "Payout Ratio: % of EPS distributed as dividends\n"
     "Irregularity Factor: fraction of years dividends are actually paid (Irregular policy only)\n"
     "Effective Payout = Stated Payout × Irregularity Factor (calculated automatically)"),
    ("Step 4", "Enter Terminal Value Assumptions (Inputs sheet)",
     "Terminal Growth Rate: the rate at which residual income grows beyond the forecast horizon.\n"
     "  → Must be below the Cost of Equity and below long-run GDP growth\n"
     "Long-Term ROE: the sustainable ROE the bank achieves in perpetuity.\n"
     "  → Must exceed Cost of Equity; otherwise the company destroys value in perpetuity"),
    ("Step 5", "Read Results (Valuation & Sensitivity sheets)",
     "Valuation sheet: shows the intrinsic value build-up (BV₀ + PV(RI) + PV(TV))\n"
     "Sensitivity Table A: intrinsic value across terminal growth × long-term ROE combinations\n"
     "Sensitivity Table B: intrinsic value across beta × terminal growth combinations\n"
     "Cover sheet: shows headline intrinsic value, market price, margin of safety, and signal"),
    ("Convention", "Input cells only",
     "Yellow cells = data entry (your inputs)\n"
     "All other cells contain formulas — do not overwrite them\n"
     "Enter all percentages as decimals (e.g. 18% → 0.18) or as whole numbers (18) depending on the cell format shown"),
]

row = 3
for step, title, detail in steps:
    c1 = instr.cell(row=row, column=1, value=step)
    style(c1, bold=True, size=10, color=WHITE, bg=TEAL, h="center", v="top")
    c2 = instr.cell(row=row, column=2, value=title)
    style(c2, bold=True, size=11, color=NAVY, bg=LBLUE)
    row += 1
    c3 = instr.cell(row=row, column=2, value=detail)
    style(c3, size=10, color="374151", wrap=True, v="top")
    instr.row_dimensions[row].height = max(60, detail.count('\n')*16+20)
    row += 2

# Reorder sheets
wb._sheets = [cover, instr, wb["ROE_Engine"], inp_ws, fc, val, sens]

# ══════════════════════════════════════════════════════════════
# CELL TIPS — Comments on all input and output cells
# ══════════════════════════════════════════════════════════════
def tip(ws, cell_ref, text):
    c = ws[cell_ref]
    cm = Comment(text, "InvestIQ")
    cm.width = 330
    cm.height = max(70, text.count('\n') * 17 + 50)
    c.comment = cm

def tip_range(ws, start_col, end_col, row, text):
    for col in range(start_col, end_col + 1):
        tip(ws, f"{get_column_letter(col)}{row}", text)

# ── Cover sheet ──────────────────────────────────────────────
tip(cover, "B10", "INTRINSIC VALUE PER SHARE\nLinked from the Valuation sheet.\nFormula: BV₀ + PV(Residual Income) + PV(Terminal Value)\n\nThis is the model’s estimate of what the stock is worth based on its ability to earn returns above the cost of equity.")
tip(cover, "B11", "Current market price per share entered in Inputs sheet (cell B28).\nLeave as 0 if you have not entered a price yet.")
tip(cover, "B12", "Margin of Safety = (Intrinsic Value − Market Price) ÷ Intrinsic Value\n\nPositive % → stock appears undervalued (price below IV).\nNegative % → stock appears overvalued (price above IV).\n\nA higher MoS gives more buffer against forecast errors.")
tip(cover, "B13", "Valuation signal based on comparing Intrinsic Value to Market Price:\n• UNDERVALUED → IV > Market Price (potential buy)\n• OVERVALUED  → IV < Market Price (potential avoid/sell)\n• FAIRLY VALUED → IV ≈ Market Price")

# ── Inputs sheet — Setup ──────────────────────────────────────
tip(inp_ws, "B4", "Select the currency for all per-share values (BV₀, EPS, Market Price).\nGHS = Ghana Cedis  |  USD = US Dollars\n\nEnsure all inputs are consistently in the same currency.")
tip(inp_ws, "B5", "Number of years to forecast explicitly before switching to the Terminal Value.\nAllowed values: 3, 4, or 5.\n\nLonger horizons capture more near-term detail but require more assumptions.\nTypical practice: 5 years for banks with stable earnings.")

# ── Inputs sheet — CAPM ──────────────────────────────────────
tip(inp_ws, "B8", "RISK-FREE RATE (Rf) — the return on a default-free investment.\n\nWhere to find: Bank of Ghana (bog.gov.gh) → Treasury Bills & Bonds page.\nUse the yield on the 5- or 10-year government bond.\n\nTypical Ghana value: 25–30% (2024–25).\nEnter as decimal: 0.28")
tip(inp_ws, "B9", "EXPECTED MARKET RETURN (Rm) — the long-run annual return of the stock market.\n\nWhere to find: GSE website (gse.com.gh) → compute the annualised GSE All-Share Index return over 5–10 years.\n\nTypical Ghana value: 20–28% annualised.\nEnter as decimal: 0.22")
tip(inp_ws, "B10", "BETA (β) — measures the stock’s volatility relative to the market.\nβ = 1.0 → moves exactly with the market.\nβ > 1.0 → more volatile (higher required return).\nβ < 1.0 → less volatile (lower required return).\n\nWhere to find: Bloomberg Terminal, Reuters Eikon, or calculate as Cov(stock, index) ÷ Var(index) over 3–5 years of weekly returns.\n\nTypical Ghana bank range: 0.5–1.2")
tip(inp_ws, "B11", "MARKET RISK PREMIUM (MRP) — auto-calculated.\nFormula: MRP = Market Return − Risk-Free Rate\n= Inputs!B9 − Inputs!B8\n\nRepresents the extra return investors demand for holding equities over risk-free bonds.")
tip(inp_ws, "B12", "COST OF EQUITY (r) — auto-calculated via CAPM.\nFormula: r = Rf + β × (Rm − Rf)\n\nThis is the hurdle rate used throughout the model:\n• To compute the annual Equity Charge\n• As the discount rate for PV of Residual Income\n• In the Terminal Value denominator (r − g)")

# ── Inputs sheet — RIM Inputs ───────────────────────────────
tip(inp_ws, "B15", "BOOK VALUE PER SHARE (BV₀) — net asset value per share at the valuation date.\n\nWhere to find: Balance Sheet → ‘Total Equity’ ÷ Number of Shares Outstanding.\nAlso available on the GSE Fact Sheet for listed companies.\n\nThis is the Year 0 starting point; all BV roll-forward begins here.\nEnter in the same currency as your EPS/Market Price.")
tip(inp_ws, "B16", "DIVIDEND POLICY — select from the dropdown:\n• No Dividends → all earnings retained; BV grows at full ROE rate\n• Regular Dividends → dividends paid every year at the stated payout ratio\n• Irregular Dividends → dividends paid in some years only; use the Irregularity Factor below to capture frequency\n\nCheck the last 5 years of annual reports to determine the correct policy.")
tip(inp_ws, "B17", "DIVIDEND PAYOUT RATIO — the fraction of earnings paid out as dividends.\n\nWhere to find: Annual Report → Directors’ Report or Dividend Announcement.\nFormula: Payout Ratio = DPS ÷ EPS\n\nTypical Ghana bank range: 30–60%.\nEnter as decimal: 0.40 (for 40%)")
tip(inp_ws, "B18", "IRREGULARITY FACTOR — only applies when Policy = ‘Irregular Dividends’.\n\nRepresents the fraction of years dividends are actually paid:\n• 1.0 = paid every year\n• 0.75 = paid 3 out of 4 years\n• 0.50 = paid every other year\n\nCheck 5+ years of dividend history in annual reports to estimate this.")
tip(inp_ws, "B19", "EFFECTIVE PAYOUT — auto-calculated; the payout rate used in the BV roll-forward.\n\nFormula:\n• No Dividends → 0%\n• Regular → Stated Payout Ratio\n• Irregular → Stated Payout × Irregularity Factor\n\nRetained Earnings each year = EPS × (1 − Effective Payout)")

# ── Inputs sheet — Terminal Value ──────────────────────────
tip(inp_ws, "B22", "TERMINAL GROWTH RATE (g) — the perpetual growth rate of residual income beyond the forecast horizon.\n\nMust be strictly BELOW the Cost of Equity (otherwise the TV formula explodes).\nConservative rule: g ≤ long-run nominal GDP growth (~5–8% for Ghana).\n\nTypical values: 3–6%.\nEnter as decimal: 0.05")
tip(inp_ws, "B23", "LONG-TERM SUSTAINABLE ROE — the ROE the company earns in perpetuity (used in Terminal Value).\n\nMust EXCEED the Cost of Equity to generate a positive Terminal Value.\nIf LT ROE ≤ r, the bank destroys value in perpetuity.\n\nBase this on management guidance, peer benchmarks, or the average of your forecast-period ROEs.\nEnter as decimal: 0.18")
tip(inp_ws, "B24", "HEALTH CHECK: Does Long-Term ROE > Cost of Equity?\n✅ YES → model will generate a positive Terminal Value (value-creating bank).\n❌ NO  → Terminal Value is zero or negative; re-examine LT ROE or CAPM inputs.")
tip(inp_ws, "B25", "HEALTH CHECK: Is Terminal Growth Rate < Cost of Equity?\n✅ YES → Terminal Value formula is valid (Gordon Growth Model holds).\n❌ NO  → Formula breaks; reduce g or increase Cost of Equity inputs.")

# ── Inputs sheet — Market Comparison ───────────────────────
tip(inp_ws, "B28", "CURRENT MARKET PRICE PER SHARE (optional).\n\nWhere to find: GSE website (gse.com.gh) or your broker’s trading platform.\n\nUsed only to compute Margin of Safety and the valuation signal.\nLeave as 0 if you are doing a standalone intrinsic value analysis.")

# ── Inputs sheet — ROE Forecast ─────────────────────────────
for col_i in range(5):
    col_ref = f"{get_column_letter(col_i + 2)}32"
    tip(inp_ws, col_ref, f"ACTIVE FORECASTED ROE — Year {col_i + 1}\nAuto-linked from ROE_Engine Active Forecast row.\n\nDo NOT type here directly. Instead:\n1. Go to the ROE_Engine sheet\n2. Select a Mode in cell B3\n3. Fill in the required inputs for that mode\n\nThe correct ROE will flow here automatically.")

# ── ROE_Engine sheet — Mode ──────────────────────────────────
tip(roe_eng, "B3", "ROE FORECAST MODE — select from the dropdown:\n• Manual        → type each year’s ROE in Section C (row 19)\n• Auto-Average  → uses simple average of historical ROEs from Section A\n• Auto-CAGR     → projects ROEs using CAGR of historical ROEs\n• Growth Rate   → compounds a base ROE at a set growth rate (Section B)\n• From-EPS      → derives ROE = Projected EPS ÷ Rolling Opening BV (Section E2)\n\nThe selected mode feeds the Active Forecast row (row 22) which flows into the Inputs sheet.")

# ── ROE_Engine — Section A ────────────────────────────────────
tip_range(roe_eng, 2, 6, 7, "HISTORICAL ROE — enter the annual Return on Equity for up to 5 past years.\n\nWhere to find:\n• Annual Report → Net Income ÷ Shareholders’ Equity\n• OR compute below in Section E (rows 28–30) from NI and Equity\n\nEnter as decimal (e.g. 18% → 0.18).\nY-5 = oldest; Y-1 = most recent year.")
tip(roe_eng, "B9", "SIMPLE AVERAGE of all historical ROEs entered in row 7.\nFormula: = AVERAGE(B7:F7)\n\nUsed as the constant forecast ROE for all years when Mode = ‘Auto-Average’.\nBest suited when ROE has been relatively stable historically.")
tip(roe_eng, "B10", "LATEST ROE — the most recent year’s ROE (Y-1 from row 7).\nFormula: = F7\n\nUsed as the baseline in Auto-CAGR trend projections.")
tip(roe_eng, "B11", "CAGR of ROE from Y-5 to Y-1 (compound annual growth rate).\nFormula: = (Latest / Oldest)^(1/4) − 1\n\nUsed when Mode = ‘Auto-CAGR’ to trend-project future ROEs.\nPositive CAGR → improving ROE trend; Negative → declining trend.")

# ── ROE_Engine — Section B ────────────────────────────────────
tip(roe_eng, "B14", "BASE ROE — the starting ROE for Growth Rate mode.\n\nTypically the most recent known ROE.\nFormula applied: Forecast ROE_t = Base ROE × (1 + Growth Rate)^t\n\nEnter as decimal: 0.18")
tip(roe_eng, "B15", "ANNUAL ROE GROWTH RATE — the rate at which ROE improves (or declines) each year.\n\nPositive → management expects improving margins/returns.\nNegative → mean-reversion or sector headwinds expected.\n\nEnter as decimal: 0.02 (for 2% annual improvement)")

# ── ROE_Engine — Section C ────────────────────────────────────
tip_range(roe_eng, 2, 6, 19, "MANUAL ROE FORECAST — enter your own ROE estimate for each year.\n\nOnly used when Mode (B3) = ‘Manual’.\n\nTypical Ghana bank range: 12–25%.\nEnter as decimal: 0.18\n\nBest when you have specific analyst forecasts or company guidance.")

# ── ROE_Engine — Section D (Active Forecast) ─────────────────
tip_range(roe_eng, 2, 6, 22, "ACTIVE FORECAST ROE — auto-selected based on Mode (B3). DO NOT EDIT.\n\nFormula chain:\n• Manual    → from Section C row 19\n• Auto-Avg  → from B9 (historical average)\n• Auto-CAGR → Latest ROE × (1 + CAGR)^t\n• Growth    → Base ROE × (1 + Growth Rate)^t\n• From-EPS  → from Section E2 row 37 (EPS ÷ Opening BV)\n\nThese values flow into the Inputs sheet ROE row (row 32).")

# ── ROE_Engine — Section E1 ───────────────────────────────────
tip_range(roe_eng, 2, 6, 28, "NET INCOME — total profit after tax for the year.\n\nWhere to find: Income Statement → last line: ‘Profit for the Year’ or ‘Net Profit After Tax’.\n\nEnter in CONSISTENT units across all years (all GHS thousands OR all GHS millions).\nY-5 = oldest year; Y-1 = most recent year.")
tip_range(roe_eng, 2, 6, 29, "SHAREHOLDERS’ EQUITY — total equity on the balance sheet.\n\nWhere to find: Balance Sheet → ‘Total Equity’ or ‘Total Shareholders’ Funds’.\nUse the OPENING balance (prior year’s closing equity) for accuracy.\n\nEnter in the SAME units as Net Income above.\nY-5 = oldest year; Y-1 = most recent year.")
tip_range(roe_eng, 2, 6, 30, "COMPUTED HISTORICAL ROE — auto-calculated.\nFormula: = Net Income ÷ Shareholders’ Equity\n\nTo use these values in the model:\n1. Copy them into Section A (row 7)\n2. Set Mode to ‘Auto-Average’ or ‘Auto-CAGR’\n\nDo not edit these cells — they are formula-driven.")

# ── ROE_Engine — Section E2 ───────────────────────────────────
tip_range(roe_eng, 2, 6, 35, "PROJECTED EPS — Earnings Per Share forecast for each future year.\n\nWhere to find:\n• Analyst research reports / broker forecasts\n• Management EPS guidance\n• Or estimate: Forecast Net Income ÷ Shares Outstanding\n\nEnter on a per-share basis in the same currency as BV₀ (Inputs!B15).\nOnly used when Mode = ‘From-EPS’.")
tip_range(roe_eng, 2, 6, 36, "OPENING BV PER SHARE — auto-calculated via roll-forward. DO NOT EDIT.\n\nFormula:\nYear 1 = BV₀ from Inputs!B15\nYear t = BV_{t-1} + EPS_{t-1} × (1 − Effective Payout)\n\nThis is the denominator in the ROE = EPS ÷ Opening BV calculation.\nEffective Payout is drawn from Inputs!B19.")
tip_range(roe_eng, 2, 6, 37, "COMPUTED FORECAST ROE — auto-calculated. DO NOT EDIT.\nFormula: = Projected EPS ÷ Opening BV per Share\n\nWhen Mode = ‘From-EPS’, these values flow automatically into the Active Forecast (row 22) above, which feeds the Inputs sheet and drives the full model.\n\nPositive ROE requires positive EPS and positive Opening BV.")

# ── Forecast sheet — row tips ──────────────────────────────────
for col_i in range(5):
    col_ref_base = get_column_letter(col_i + 2)
    tip(fc, f"{col_ref_base}4",  f"Shows whether Year {col_i+1} is within the chosen forecast horizon.\nFormula: = IF({col_i+1} ≤ Inputs!B5, ‘Yes’, ‘—’)\n‘—’ means this year is beyond the horizon and all its values are blank.")
    tip(fc, f"{col_ref_base}5",  f"FORECASTED ROE for Year {col_i+1} — pulled from Inputs sheet (linked to ROE_Engine).\nFormula: = IF(active, Inputs!{col_ref_base}32, ‘’)\n\nTo change this ROE, go to ROE_Engine and adjust the Mode or inputs there.")
    tip(fc, f"{col_ref_base}6",  f"OPENING BOOK VALUE PER SHARE — start of Year {col_i+1}.\n{'Formula: = BV₀ from Inputs!B15 (Year 0 starting point).' if col_i==0 else f'Formula: = Closing BV of Year {col_i} (previous row 10).'}\n\nThis is the equity base on which ROE is applied.")
    tip(fc, f"{col_ref_base}7",  f"NET INCOME PER SHARE (model equivalent of EPS) for Year {col_i+1}.\nFormula: = Forecasted ROE × Opening BV\n\nRepresents the earnings generated on the opening equity base.")
    tip(fc, f"{col_ref_base}8",  f"DIVIDEND PER SHARE for Year {col_i+1}.\nFormula: = Net Income per Share × Effective Payout (Inputs!B19)\n\nAffected by Dividend Policy and Irregularity Factor set in the Inputs sheet.")
    tip(fc, f"{col_ref_base}9",  f"RETAINED EARNINGS PER SHARE for Year {col_i+1}.\nFormula: = Net Income per Share − Dividend per Share\n\nThis amount is added to the Book Value, driving BV growth.")
    tip(fc, f"{col_ref_base}10", f"CLOSING BOOK VALUE PER SHARE — end of Year {col_i+1}.\nFormula: = Opening BV + Retained Earnings\n\nBecomes the Opening BV for Year {col_i+2} (feeds next year’s row 6).")
    tip(fc, f"{col_ref_base}11", f"EQUITY CHARGE for Year {col_i+1} — the minimum return required on the opening equity.\nFormula: = Cost of Equity (r) × Opening BV\n\nThis is the ‘hurdle’: if Net Income doesn’t exceed this, Residual Income is negative.")
    tip(fc, f"{col_ref_base}12", f"RESIDUAL INCOME for Year {col_i+1} — the ‘super-profit’ above the equity hurdle.\nFormula: = Net Income per Share − Equity Charge\n\nPositive RI → value creation (ROE > r).\nNegative RI → value destruction (ROE < r).")
    tip(fc, f"{col_ref_base}13", f"DISCOUNT FACTOR for Year {col_i+1} — present value of £1 received in Year {col_i+1}.\nFormula: = 1 ÷ (1 + r)^{col_i+1}\n\nA higher Cost of Equity produces a smaller discount factor and lower present values.")
    tip(fc, f"{col_ref_base}14", f"PRESENT VALUE OF RESIDUAL INCOME for Year {col_i+1}.\nFormula: = Residual Income × Discount Factor\n\nThese are summed across all active years for the PV(RI) component of Intrinsic Value.")

# ── Valuation sheet ───────────────────────────────────────────
tip(val, "B4",  "Cost of Equity (r) — linked from Inputs!B12.\nCAPM formula: r = Rf + β × (Rm − Rf)\nUsed as both the discount rate and the equity charge rate throughout the model.")
tip(val, "B5",  "Forecast horizon — number of explicit forecast years (3, 4, or 5).\nLinked from Inputs!B5.\nDetermines how many columns of the Forecast sheet are active.")
tip(val, "B8",  "BV₀ — Current Book Value per Share (starting point of the valuation).\nLinked from Inputs!B15.\n\nThis is the ‘floor’ value — the minimum IV before any residual income or terminal value is added.\nHigher-quality banks with strong RI add significantly above this floor.")
tip(val, "B9",  "PV OF RESIDUAL INCOME — sum of discounted residual incomes across the explicit forecast period.\nFormula: = SUM(Forecast!B14:F14)\n\nCaptures the value created during the years the model explicitly forecasts.\nIf all ROEs ≤ r, this sum will be zero or negative.")
tip(val, "B10", "TERMINAL BOOK VALUE PER SHARE — BV at the END of the final forecast year.\nLinked from Forecast row 10 (last active year).\n\nUsed as the equity base for computing Terminal Residual Income.\nHigher retained earnings (lower payout) → higher Terminal BV → higher Terminal Value.")
tip(val, "B11", "TERMINAL RESIDUAL INCOME — excess earnings in the first year beyond the forecast horizon.\nFormula: = (Long-Term ROE − r) × Terminal BV\n\nMust be positive (Long-Term ROE > r) for the Terminal Value to contribute positively to IV.")
tip(val, "B12", "TERMINAL VALUE — present value of all residual income beyond the forecast horizon, assumed to grow at rate g perpetually.\nFormula: = Terminal RI × (1 + g) ÷ (r − g)\n\nThis is usually the largest single component of IV for a growing bank.\nSee Sensitivity sheet for how IV changes with g and LT ROE.")
tip(val, "B13", "PV OF TERMINAL VALUE — Terminal Value discounted back to today.\nFormula: = Terminal Value ÷ (1 + r)^n  where n = forecast years\n\nLonger forecast horizons discount TV more heavily, reducing its current contribution to IV.")
tip(val, "B14", "INTRINSIC VALUE PER SHARE\nFormula: = BV₀ + PV(Residual Income) + PV(Terminal Value)\n\nThis is the model’s estimate of the stock’s fair value. Compare to Market Price (B17) to assess whether the stock is under- or over-valued.\n\nSensitivity to inputs: see the Sensitivity sheet for how IV changes across key assumption ranges.")
tip(val, "B17", "Current market price per share — entered in Inputs!B28.\nUsed only for Market Comparison. Does not affect the Intrinsic Value calculation.")
tip(val, "B18", "UPSIDE (DOWNSIDE) = Intrinsic Value − Market Price.\nPositive → stock trades below IV (potential upside).\nNegative → stock trades above IV (potential downside).")
tip(val, "B19", "MARGIN OF SAFETY = (IV − Market Price) ÷ IV\n\nHigher MoS → greater discount to intrinsic value → more buffer against forecast errors.\nA common threshold: >20–25% MoS before considering an investment.")
tip(val, "B20", "PRICE-TO-BOOK RATIO = Market Price ÷ Book Value per Share.\n\nIndicates how much the market pays for each unit of book equity.\n>1.0 → market expects the bank to earn above its cost of equity.\n<1.0 → market implies ROE below cost of equity.")
tip(val, "B21", "INTRINSIC VALUE-TO-BOOK RATIO = IV ÷ BV₀.\n\nShows how much of the IV is above book value (driven by excess returns).\nCompare to P/BV: if IV/BV > P/BV, model implies the stock is undervalued.")
tip(val, "B22", "VALUATION SIGNAL — simple directional indicator.\n• UNDERVALUED → IV > Market Price (model says potential buy)\n• OVERVALUED  → IV < Market Price (model says potential sell/avoid)\n\nAlways use with judgment — the signal is only as good as the input assumptions.")

# Sensitivity sheet header tips
tip(sens, "B3",  "TABLE A: Intrinsic Value across combinations of Terminal Growth Rate and Long-Term ROE.\n\nRows vary the terminal growth rate (g) around your base assumption.\nColumns vary the Long-Term ROE around your base assumption.\n\nThe highlighted cell (centre) is your base-case IV.")
tip(sens, "B12", "TABLE B: Intrinsic Value across combinations of Beta and Terminal Growth Rate.\n\nRows vary Beta (β) around your base input, which changes the Cost of Equity.\nColumns vary the terminal growth rate (g) around your base assumption.\n\n‘N/A’ appears where g ≥ r (model undefined) or LT ROE ≤ r (no value creation).")

wb.save(OUT)
print(f"Saved: {OUT}")
